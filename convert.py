#!/usr/bin/env python3
"""Convert DSA_500_Problems_Roadmap.xlsx to data.json for the tracker website."""

import json
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("DSA_500_Problems_Roadmap.xlsx")

# --- Sheet 1: 500 DSA Problems ---
ws = wb["500 DSA Problems"]
problems = []
all_topics = set()
all_difficulties = set()
all_priorities = set()
all_patterns = set()
total_easy = total_medium = total_hard = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    num, topic, lc_num, name, difficulty, pattern, priority, link, insight, status, date_solved, notes = row
    if num is None or name is None:
        continue

    topic = str(topic).strip() if topic else ""
    difficulty = str(difficulty).strip() if difficulty else ""
    pattern = str(pattern).strip() if pattern else ""
    priority = str(priority).strip() if priority else ""

    all_topics.add(topic)
    all_difficulties.add(difficulty)
    if priority:
        all_priorities.add(priority)
    if pattern:
        all_patterns.add(pattern)

    if difficulty == "Easy":
        total_easy += 1
    elif difficulty == "Medium":
        total_medium += 1
    elif difficulty == "Hard":
        total_hard += 1

    solved = "solved" if status and str(status).strip() not in ("⬜", "") else "unsolved"

    problems.append({
        "id": int(num),
        "topic": topic,
        "lcNumber": str(lc_num).strip() if lc_num else "",
        "name": str(name).strip(),
        "difficulty": difficulty,
        "pattern": pattern,
        "priority": priority,
        "link": str(link).strip() if link else "",
        "keyInsight": str(insight).strip() if insight else "",
        "status": solved,
        "dateSolved": str(date_solved) if date_solved else None,
        "notes": str(notes).strip() if notes else ""
    })

# --- Sheet 2: 6-Month Daily Plan ---
ws2 = wb["6-Month Daily Plan"]
study_plan = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    week, day, date_range, topic_focus, probs, goal, status = row
    if week is None and day is None:
        continue
    study_plan.append({
        "week": str(week).strip() if week else "",
        "day": str(day).strip() if day else "",
        "dateRange": str(date_range).strip() if date_range else "",
        "topicFocus": str(topic_focus).strip() if topic_focus else "",
        "problems": str(probs).strip() if probs else "",
        "goal": str(goal).strip() if goal else "",
        "status": "completed" if status and str(status).strip() not in ("⬜", "") else "pending"
    })

# --- Sheet 3: Topic Summary ---
ws3 = wb["Topic Summary"]
topic_summary = []
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True):
    topic, total, easy, medium, hard, patterns = row
    if topic is None or str(topic).strip() == "TOTAL":
        continue
    topic_summary.append({
        "topic": str(topic).strip(),
        "total": int(total) if total else 0,
        "easy": int(easy) if easy else 0,
        "medium": int(medium) if medium else 0,
        "hard": int(hard) if hard else 0,
        "keyPatterns": str(patterns).strip() if patterns else ""
    })

# --- Build output ---
data = {
    "problems": problems,
    "studyPlan": study_plan,
    "topicSummary": topic_summary,
    "metadata": {
        "totalProblems": len(problems),
        "totalEasy": total_easy,
        "totalMedium": total_medium,
        "totalHard": total_hard,
        "topics": sorted(all_topics),
        "difficulties": ["Easy", "Medium", "Hard"],
        "priorities": sorted(all_priorities),
        "patterns": sorted(all_patterns),
        "generatedAt": datetime.now().isoformat()
    }
}

with open("data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Generated data.json:")
print(f"  Problems: {len(problems)}")
print(f"  Study Plan entries: {len(study_plan)}")
print(f"  Topic summaries: {len(topic_summary)}")
print(f"  Difficulties: Easy={total_easy}, Medium={total_medium}, Hard={total_hard}")
print(f"  Topics: {len(all_topics)}")
print(f"  Patterns: {len(all_patterns)}")
